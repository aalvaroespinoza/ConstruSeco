import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from datetime import datetime

from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QMessageBox, QFileDialog, QTableWidget, QTableWidgetItem, QHeaderView,
    QProgressBar, QScrollArea, QWidget
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor

from db.queries import obtener_stocks_todos

COLOR_PRIMARY = "#2563eb"
COLOR_BG = "#f8fafc"
COLOR_CARD_BG = "#ffffff"
COLOR_BORDER = "#e2e8f0"
COLOR_TEXT_MAIN = "#1e293b"
COLOR_SUCCESS = "#10b981"
COLOR_DANGER = "#ef4444"


def normalizar_unidad(u: str) -> str:
    if not u:
        return 'u'
    ul = str(u).strip().lower()
    if ul in ['u', 'unidad', 'unidades']:
        return 'u'
    if ul in ['m2', 'm²']:
        return 'm2'
    return None


def generar_plantilla_excel(parent_window):
    ruta, _ = QFileDialog.getSaveFileName(
        parent_window, "Guardar Plantilla de Importación",
        "plantilla_importacion_stock.xlsx", "Excel Files (*.xlsx)"
    )
    if not ruta:
        return

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Importación de Productos"
        
        columnas = ["codigo", "descripcion", "unidad", "precio_venta", "stock_inicial", "stock_minimo"]
        ws.append(columnas)
        
        header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
            
        # Fila de ejemplo
        ws.append(["PROD-001", "Cemento Portland 50kg", "u", 8500, 100, 10])
        
        wb.save(ruta)
        QMessageBox.information(parent_window, "Éxito", "Plantilla generada correctamente.")
    except Exception as e:
        QMessageBox.critical(parent_window, "Error", f"No se pudo generar la plantilla: {e}")


def exportar_inventario_excel(conn, parent_window):
    ruta, _ = QFileDialog.getSaveFileName(
        parent_window, "Exportar Inventario",
        f"Inventario_{datetime.now().strftime('%Y%m%d')}.xlsx", "Excel Files (*.xlsx)"
    )
    if not ruta:
        return

    try:
        stocks = obtener_stocks_todos(conn)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario"
        
        columnas = [
            "Código", "Descripción", "Unidad", "Stock Físico", "Comprometido", 
            "Disponible", "Stock Mínimo", "Precio Venta", "Valor Inventario", "Estado"
        ]
        ws.append(columnas)
        
        header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
        
        for p in stocks:
            atp = p["atp"]
            stk_min = p["stock_minimo"]
            estado = "Disponible"
            if atp <= 0: estado = "Sin Stock"
            elif stk_min > 0 and atp <= stk_min: estado = "Stock Bajo"
            
            valor_inv = p["stock_fisico"] * p["precio_venta"]
            
            ws.append([
                p["codigo"],
                p["descripcion"],
                p["unidad_base"],
                p["stock_fisico"],
                p["comprometido"],
                atp,
                stk_min,
                p["precio_venta"],
                valor_inv,
                estado
            ])
            
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2
            
        wb.save(ruta)
        QMessageBox.information(parent_window, "Éxito", f"Inventario exportado correctamente a:\n{ruta}")
    except Exception as e:
        QMessageBox.critical(parent_window, "Error", f"No se pudo exportar el inventario: {e}")


class DialogoImportarExcel(QDialog):
    def __init__(self, conexion_db, parent=None):
        super().__init__(parent)
        self.conn = conexion_db
        self.setWindowTitle("Importar Productos desde Excel")
        self.setMinimumWidth(800)
        self.setMinimumHeight(600)
        self.filas_validas = []
        self.errores = []
        self.columnas_map = {}
        self.modo_importacion = "AGREGAR"
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # Opciones de importacion
        ly_modos = QHBoxLayout()
        ly_modos.addWidget(QLabel("Modo de Importación:"))
        
        from PyQt6.QtWidgets import QRadioButton
        self.rb_agregar = QRadioButton("Agregar / Actualizar Productos (Seguro)")
        self.rb_agregar.setChecked(True)
        self.rb_agregar.toggled.connect(self.cambiar_modo)
        
        self.rb_sustituir = QRadioButton("Sustituir Catálogo Actual (Peligroso)")
        self.rb_sustituir.setStyleSheet("color: #ef4444; font-weight: bold;")
        self.rb_sustituir.toggled.connect(self.cambiar_modo)
        
        ly_modos.addWidget(self.rb_agregar)
        ly_modos.addWidget(self.rb_sustituir)
        ly_modos.addStretch()
        layout.addLayout(ly_modos)
        
        self.lbl_advertencia = QLabel("ADVERTENCIA: esta operación puede afectar todo el catálogo actual y no debe ejecutarse sin una copia de seguridad.")
        self.lbl_advertencia.setStyleSheet("color: white; background-color: #ef4444; padding: 8px; border-radius: 4px; font-weight: bold;")
        self.lbl_advertencia.hide()
        layout.addWidget(self.lbl_advertencia)
        
        # Panel superior
        ly_top = QHBoxLayout()
        self.lbl_archivo = QLabel("Ningún archivo seleccionado")
        self.lbl_archivo.setStyleSheet(f"color: {COLOR_TEXT_MAIN}; background-color: {COLOR_CARD_BG}; padding: 8px; border: 1px solid {COLOR_BORDER}; border-radius: 4px;")
        
        btn_seleccionar = QPushButton("📂 Seleccionar Excel")
        btn_seleccionar.setStyleSheet(f"background-color: {COLOR_PRIMARY}; color: white; padding: 8px 16px; border-radius: 4px; font-weight: bold;")
        btn_seleccionar.clicked.connect(self.seleccionar_archivo)
        
        ly_top.addWidget(self.lbl_archivo, stretch=1)
        ly_top.addWidget(btn_seleccionar)
        layout.addLayout(ly_top)
        
        # Resumen
        self.lbl_resumen = QLabel("")
        self.lbl_resumen.setStyleSheet(f"font-weight: bold; margin-top: 10px;")
        layout.addWidget(self.lbl_resumen)
        
        # Tabla de errores
        self.tabla_errores = QTableWidget(0, 2)
        self.tabla_errores.setHorizontalHeaderLabels(["Fila (Excel)", "Error detectado"])
        self.tabla_errores.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.tabla_errores.setStyleSheet(f"border: 1px solid {COLOR_DANGER}; background-color: #fef2f2;")
        self.tabla_errores.hide()
        layout.addWidget(self.tabla_errores)
        
        # Tabla de correctos (previsualizacion)
        self.tabla_preview = QTableWidget(0, 6)
        self.tabla_preview.setHorizontalHeaderLabels(["Código", "Descripción", "Unidad", "Precio", "Stk Inicial", "Stk Mínimo"])
        self.tabla_preview.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.tabla_preview.setStyleSheet(f"border: 1px solid {COLOR_SUCCESS}; background-color: #f0fdf4;")
        self.tabla_preview.hide()
        layout.addWidget(self.tabla_preview)
        
        # Boton Confirmar
        self.btn_confirmar = QPushButton("✅ Confirmar Importación")
        self.btn_confirmar.setStyleSheet(f"background-color: {COLOR_SUCCESS}; color: white; padding: 12px; border-radius: 4px; font-weight: bold; font-size: 14px;")
        self.btn_confirmar.clicked.connect(self.ejecutar_importacion)
        self.btn_confirmar.setEnabled(False)
        layout.addWidget(self.btn_confirmar)

    def cambiar_modo(self):
        if self.rb_sustituir.isChecked():
            self.modo_importacion = "SUSTITUIR"
            self.lbl_advertencia.show()
            self.btn_confirmar.setStyleSheet(f"background-color: {COLOR_DANGER}; color: white; padding: 12px; border-radius: 4px; font-weight: bold; font-size: 14px;")
        else:
            self.modo_importacion = "AGREGAR"
            self.lbl_advertencia.hide()
            self.btn_confirmar.setStyleSheet(f"background-color: {COLOR_SUCCESS}; color: white; padding: 12px; border-radius: 4px; font-weight: bold; font-size: 14px;")
            
        if self.lbl_archivo.text() != "Ningún archivo seleccionado":
            self.procesar_archivo(self.lbl_archivo.text())

    def seleccionar_archivo(self):
        ruta, _ = QFileDialog.getOpenFileName(self, "Seleccionar Excel de Importación", "", "Excel Files (*.xlsx)")
        if not ruta:
            return
            
        self.lbl_archivo.setText(ruta)
        self.procesar_archivo(ruta)

    def auto_mapear_columnas(self, headers):
        mapa = {'codigo': -1, 'descripcion': -1, 'unidad': -1, 'precio_venta': -1, 'stock_inicial': -1, 'stock_minimo': -1}
        aliases = {
            'codigo': ['código', 'codigo', 'sku', 'sku / código', 'código de barras'],
            'descripcion': ['producto', 'descripción', 'descripcion', 'articulo', 'artículo', 'material', 'artículo / material'],
            'unidad': ['unidad', 'medida', 'unidad de medida'],
            'stock_inicial': ['stock', 'existencia', 'existencia actual', 'cantidad', 'stock físico', 'stock inicial'],
            'precio_venta': ['precio', 'precio venta', 'precio unitario', 'precio_venta'],
            'stock_minimo': ['stock mínimo', 'mínimo', 'mínimo deseado', 'stock minimo', 'stock_minimo']
        }
        
        for idx, h in enumerate(headers):
            if not h: continue
            hl = str(h).strip().lower()
            for key, val_list in aliases.items():
                if hl in val_list and mapa[key] == -1:
                    mapa[key] = idx
                    break
        return mapa

    def procesar_archivo(self, ruta):
        self.filas_validas = []
        self.errores = []
        self.btn_confirmar.setEnabled(False)
        self.tabla_errores.setRowCount(0)
        self.tabla_preview.setRowCount(0)
        
        try:
            wb = openpyxl.load_workbook(ruta, data_only=True)
            ws = wb.active
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo leer el archivo Excel: {e}")
            return
            
        filas = list(ws.iter_rows(values_only=True))
        if not filas or len(filas) < 2:
            QMessageBox.warning(self, "Error", "El archivo está vacío o no tiene datos válidos.")
            return
            
        headers = filas[0]
        mapa = self.auto_mapear_columnas(headers)
        
        # Validar mapeo
        if mapa['codigo'] == -1 or mapa['descripcion'] == -1:
            msg = (
                "No se encontraron las columnas obligatorias.\n\n"
                "Para poder importar, la primera fila de tu Excel DEBE contener encabezados reconocibles. "
                "Las columnas requeridas son:\n\n"
                "1. Código del producto (nombres aceptados: código, sku, código de barras)\n"
                "2. Descripción del producto (nombres aceptados: producto, descripción, artículo, material)\n\n"
                "Opcionalmente también podés incluir:\n"
                "- Unidad (nombres: unidad, medida)\n"
                "- Precio (nombres: precio, precio venta)\n"
                "- Stock Inicial (nombres: stock, existencia, cantidad)\n"
                "- Stock Mínimo (nombres: stock mínimo, mínimo)\n\n"
                "Por favor, revisá tu archivo, asegurate de que los nombres estén en la primera fila, guardá y volvé a intentarlo."
            )
            QMessageBox.critical(self, "Formato No Reconocido", msg)
            return
            
        c = self.conn.cursor()
        c.execute("SELECT codigo, lower(descripcion) FROM productos WHERE activo = 1")
        existentes = c.fetchall()
        codigos_db = set(str(r[0]).strip() for r in existentes)
        desc_db = set(str(r[1]).strip() for r in existentes)
        
        codigos_excel = set()
        desc_excel = set()
        
        for i, fila in enumerate(filas[1:], start=2):
            if not any(fila): continue
            
            try:
                cod = str(fila[mapa['codigo']]).strip() if mapa['codigo'] != -1 and fila[mapa['codigo']] is not None else ""
                desc = str(fila[mapa['descripcion']]).strip() if mapa['descripcion'] != -1 and fila[mapa['descripcion']] is not None else ""
                uni_raw = str(fila[mapa['unidad']]).strip() if mapa['unidad'] != -1 and fila[mapa['unidad']] is not None else ""
                precio = float(fila[mapa['precio_venta']]) if mapa['precio_venta'] != -1 and fila[mapa['precio_venta']] is not None else 0.0
                stk_ini = float(fila[mapa['stock_inicial']]) if mapa['stock_inicial'] != -1 and fila[mapa['stock_inicial']] is not None else 0.0
                stk_min = float(fila[mapa['stock_minimo']]) if mapa['stock_minimo'] != -1 and fila[mapa['stock_minimo']] is not None else 0.0
            except ValueError:
                self.errores.append((i, "Valores numéricos inválidos en la fila."))
                continue
                
            errs = []
            if not cod: errs.append("Código vacío.")
            if not desc: errs.append("Descripción vacía.")
            if precio < 0 or stk_ini < 0 or stk_min < 0: errs.append("Valores negativos no permitidos.")
            
            uni = normalizar_unidad(uni_raw)
            if not uni: errs.append(f"Unidad '{uni_raw}' inválida. Use 'u' o 'm2'.")
            
            if self.modo_importacion == "AGREGAR":
                if cod in codigos_excel: errs.append("Código duplicado dentro del Excel.")
            else:
                if cod in codigos_excel: errs.append("Código duplicado dentro del Excel.")
                
            if errs:
                self.errores.append((i, " | ".join(errs)))
            else:
                codigos_excel.add(cod)
                desc_excel.add(desc.lower())
                
                estado_update = "NUEVO"
                if cod in codigos_db:
                    estado_update = "ACTUALIZAR"
                    
                self.filas_validas.append((cod, desc, uni, precio, stk_ini, stk_min, estado_update))
                
        if self.errores:
            self.filas_validas.clear()
            
        self.mostrar_resultados()

    def mostrar_resultados(self):
        total_validas = len(self.filas_validas)
        total_errores = len(self.errores)
        
        if total_errores > 0:
            self.lbl_resumen.setText(f"❌ Se encontraron {total_errores} fila(s) con errores. Corrige el archivo y vuelve a intentarlo.")
            self.lbl_resumen.setStyleSheet(f"color: {COLOR_DANGER}; font-weight: bold; font-size: 14px;")
            self.tabla_errores.show()
            self.tabla_preview.hide()
            
            self.tabla_errores.setRowCount(total_errores)
            for idx, (fila_num, err_txt) in enumerate(self.errores):
                self.tabla_errores.setItem(idx, 0, QTableWidgetItem(f"Fila {fila_num}"))
                self.tabla_errores.setItem(idx, 1, QTableWidgetItem(err_txt))
        else:
            self.lbl_resumen.setText(f"✅ Todo correcto. Se importarán/actualizarán {total_validas} producto(s).")
            self.lbl_resumen.setStyleSheet(f"color: {COLOR_SUCCESS}; font-weight: bold; font-size: 14px;")
            self.tabla_errores.hide()
            self.tabla_preview.show()
            
            self.tabla_preview.setRowCount(total_validas)
            for idx, prod in enumerate(self.filas_validas):
                for col_idx in range(6):
                    it = QTableWidgetItem(str(prod[col_idx]))
                    if prod[6] == "ACTUALIZAR":
                        it.setBackground(QColor("#fef08a")) # Amarillo claro para updates
                    self.tabla_preview.setItem(idx, col_idx, it)
                    
            self.btn_confirmar.setEnabled(True)

    def ejecutar_importacion(self):
        if not self.filas_validas: return
        
        if self.modo_importacion == "SUSTITUIR":
            reply = QMessageBox.warning(
                self, "Confirmar Sustitución",
                "Está a punto de reemplazar el catálogo actual. Los productos que no estén en el Excel y tengan historial serán desactivados, los que no, eliminados.\n\n¿Desea continuar?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                return
        
        c = self.conn.cursor()
        try:
            c.execute("BEGIN TRANSACTION;")
            fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ts = int(datetime.now().timestamp())
            
            if self.modo_importacion == "SUSTITUIR":
                # Auditar y desactivar/eliminar actuales
                c.execute("SELECT codigo FROM productos")
                todos_actuales = [r[0] for r in c.fetchall()]
                nuevos_codigos = set(p[0] for p in self.filas_validas)
                
                for cod in todos_actuales:
                    if cod not in nuevos_codigos:
                        c.execute("SELECT COUNT(*) FROM detalle_documentos WHERE codigo_producto=?", (cod,))
                        c_doc = c.fetchone()[0]
                        c.execute("SELECT COUNT(*) FROM movimientos_stock WHERE codigo_producto=?", (cod,))
                        c_mov = c.fetchone()[0]
                        c.execute("SELECT COUNT(*) FROM compromisos_stock WHERE codigo_producto=?", (cod,))
                        c_comp = c.fetchone()[0]
                        
                        if c_doc > 0 or c_mov > 0 or c_comp > 0:
                            c.execute("UPDATE productos SET activo = 0 WHERE codigo = ?", (cod,))
                        else:
                            c.execute("DELETE FROM productos WHERE codigo = ?", (cod,))
            
            for (cod, desc, uni, precio, stk_ini, stk_min, estado_update) in self.filas_validas:
                if estado_update == "ACTUALIZAR" and self.modo_importacion != "SUSTITUIR":
                    c.execute("""
                        UPDATE productos SET descripcion=?, unidad_base=?, precio_venta=?, stock_minimo=?, activo=1
                        WHERE codigo=?
                    """, (desc, uni, precio, stk_min, cod))
                    # NO agregamos stock inicial para actualizaciones en modo AGREGAR para evitar duplicar
                else:
                    # Insertar o actualizar
                    c.execute("SELECT COUNT(*) FROM productos WHERE codigo=?", (cod,))
                    if c.fetchone()[0] > 0:
                        c.execute("""
                            UPDATE productos SET descripcion=?, unidad_base=?, precio_venta=?, stock_minimo=?, activo=1
                            WHERE codigo=?
                        """, (desc, uni, precio, stk_min, cod))
                    else:
                        c.execute("""
                            INSERT INTO productos (codigo, descripcion, unidad_base, precio_venta, stock_minimo, activo)
                            VALUES (?, ?, ?, ?, ?, 1)
                        """, (cod, desc, uni, precio, stk_min))
                    
                    if stk_ini > 0:
                        c.execute("""
                            INSERT INTO documentos (numero_interno, tipo, estado, fecha_emision, observaciones)
                            VALUES (?, 'AJUSTE', 'CONFIRMADO', ?, 'Importación Excel (Stock inicial)')
                        """, (f"IMP-{cod}-{ts}", fecha))
                        id_doc = c.lastrowid
                        
                        c.execute("""
                            INSERT INTO movimientos_stock (codigo_producto, tipo_movimiento, cantidad, id_documento_origen, fecha_hora, notas)
                            VALUES (?, 'ENTRADA', ?, ?, ?, 'Importación Excel')
                        """, (cod, stk_ini, id_doc, fecha))
            
            self.conn.commit()
            QMessageBox.information(self, "Éxito", f"Se procesaron {len(self.filas_validas)} productos exitosamente.")
            self.accept()
        except Exception as e:
            self.conn.rollback()
            QMessageBox.critical(self, "Error Fatal", f"La importación falló y se revirtió. Detalles:\n{e}")
