from ui.core.modal import DialogoModalIntegrado
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime

from PyQt6.QtWidgets import (
    QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QMessageBox, QFileDialog, QTableWidget, QTableWidgetItem, QHeaderView
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor

from db.queries import obtener_stocks_todos
from ui.core.theme import (
    COLOR_PRIMARY, COLOR_CARD_BG,
    COLOR_BORDER, COLOR_TEXT_MAIN, COLOR_SUCCESS, COLOR_DANGER
)

def normalizar_unidad(u: str) -> str:
    if not u:
        return 'u'
    ul = str(u).strip().lower()
    if ul in ['u', 'unidad', 'unidades']:
        return 'u'
    if ul in ['m2', 'm²']:
        return 'm2'
    return None

def generar_plantilla_excel(parent_window):
    ruta, _ = QFileDialog.getSaveFileName(
        parent_window, "Guardar Plantilla de Importación",
        "plantilla_importacion_stock.xlsx", "Excel Files (*.xlsx)"
    )
    if not ruta:
        return

    try:
        wb = openpyxl.Workbook()
        
        # 1. Hoja principal
        ws = wb.active
        ws.title = "Plantilla"
        
        columnas = ["Código (*)", "Descripción (*)", "Unidad (*)", "Precio Venta (*)", "Stock Inicial", "Stock Mínimo", "Imagen"]
        ws.append(columnas)
        
        header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_num)].width = 25
            
        # Freeze y Filtro
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:G1"
            
        # Filas de ejemplo
        ws.append(["PROD-001", "Cemento Portland 50kg", "u", 8500, 100, 10, ""])
        ws.append(["PROD-002", "Porcelanato Beige 60x60", "m2", 15000, 50, 5, ""])
        
        # 2. Hoja de instrucciones
        ws_inst = wb.create_sheet(title="INSTRUCCIONES")
        
        instrucciones = [
            ["INSTRUCCIONES PARA LA IMPORTACIÓN MASIVA", ""],
            ["", ""],
            ["1. Campos Obligatorios (*)", "Código, Descripción, Unidad y Precio Venta son obligatorios para productos nuevos."],
            ["2. Formato Esperado", "Unidad debe ser 'u' o 'm2'. El precio y stock deben ser números positivos."],
            ["3. Códigos Duplicados en Excel", "Si el archivo Excel contiene el mismo código en varias filas, se marcarán como error."],
            ["4. Productos Existentes", "Si el código ya existe en el sistema, SOLO se actualizarán sus datos maestros (Descripción, Unidad, Precio, Mínimo, Imagen)."],
            ["5. Stock Inicial (Existentes)", "Para productos EXISTENTES, la columna Stock Inicial SERÁ IGNORADA para proteger la trazabilidad y no alterar el historial real de movimientos."],
            ["6. Stock Inicial (Nuevos)", "Para productos NUEVOS, un Stock Inicial mayor a 0 generará automáticamente un movimiento de ENTRADA (Ajuste Inicial) asociando el stock correctamente al sistema ATP."],
            ["7. Imagen", "Nombre del archivo de imagen (ej: cemento.jpg) guardado en la carpeta de imágenes, o dejar vacío."]
        ]
        
        for row in instrucciones:
            ws_inst.append(row)
            
        ws_inst.column_dimensions["A"].width = 30
        ws_inst.column_dimensions["B"].width = 120
        
        for row in ws_inst.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True, size=14)
                
        for row in ws_inst.iter_rows(min_row=3, max_row=len(instrucciones)):
            row[0].font = Font(bold=True)
            row[1].alignment = Alignment(wrap_text=True)

        wb.save(ruta)
        QMessageBox.information(parent_window, "Éxito", "Plantilla con instrucciones generada correctamente.")
    except Exception as e:
        QMessageBox.critical(parent_window, "Error", f"No se pudo generar la plantilla: {e}")

def exportar_inventario_excel(conn, parent_window):
    ruta, _ = QFileDialog.getSaveFileName(
        parent_window, "Exportar Inventario",
        f"Inventario_{datetime.now().strftime('%Y%m%d')}.xlsx", "Excel Files (*.xlsx)"
    )
    if not ruta:
        return

    try:
        stocks = obtener_stocks_todos(conn)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario"
        
        columnas = [
            "Código", "Producto / Descripción", "Unidad", "Precio de Venta", 
            "Stock Físico", "Stock Comprometido", "Stock Disponible (ATP)", 
            "Stock Mínimo", "Estado"
        ]
        ws.append(columnas)
        
        header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
        
        for p in stocks:
            atp = p["atp"]
            stk_min = p["stock_minimo"]
            estado = "Disponible"
            if atp <= 0: estado = "Sin Stock"
            elif stk_min > 0 and atp <= stk_min: estado = "Stock Bajo"
            
            row = [
                p["codigo"],
                p["descripcion"],
                p["unidad_base"],
                p["precio_venta"],
                p["stock_fisico"],
                p["comprometido"],
                atp,
                stk_min,
                estado
            ]
            ws.append(row)
            
        # Formateo
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:I{len(stocks)+1}"
        
        for row in ws.iter_rows(min_row=2, max_row=len(stocks)+1):
            row[3].number_format = '"$"#,##0.00' # Precio Venta
            row[4].number_format = '#,##0.00' # Fisico
            row[5].number_format = '#,##0.00' # Comprometido
            row[6].number_format = '#,##0.00' # ATP
            row[7].number_format = '#,##0.00' # Minimo
            
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
            
        wb.save(ruta)
        QMessageBox.information(parent_window, "Éxito", f"Inventario exportado correctamente a:\n{ruta}")
    except Exception as e:
        QMessageBox.critical(parent_window, "Error", f"No se pudo exportar el inventario: {e}")

class DialogoImportarExcel(DialogoModalIntegrado):
    def __init__(self, conexion_db, parent=None):
        super().__init__(parent)
        self.conn = conexion_db
        self.setWindowTitle("Importar Productos desde Excel")
        self.setMinimumWidth(900)
        self.setMinimumHeight(600)
        self.filas_procesadas = []
        self.columnas_map = {}
        self.stats = {'total': 0, 'nuevos': 0, 'existentes': 0, 'errores': 0, 'ignoradas': 0}
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # Panel superior
        ly_top = QHBoxLayout()
        self.lbl_archivo = QLabel("Ningún archivo seleccionado")
        self.lbl_archivo.setStyleSheet(f"color: {COLOR_TEXT_MAIN}; background-color: {COLOR_CARD_BG}; padding: 8px; border: 1px solid {COLOR_BORDER}; border-radius: 4px;")
        
        btn_seleccionar = QPushButton("📂 Seleccionar Excel")
        btn_seleccionar.setStyleSheet(f"background-color: {COLOR_PRIMARY}; color: white; padding: 8px 16px; border-radius: 4px; font-weight: bold;")
        btn_seleccionar.clicked.connect(self.seleccionar_archivo)
        
        ly_top.addWidget(self.lbl_archivo, stretch=1)
        ly_top.addWidget(btn_seleccionar)
        layout.addLayout(ly_top)
        
        # Resumen
        self.lbl_resumen = QLabel("Esperando archivo...")
        self.lbl_resumen.setStyleSheet("font-weight: bold; margin-top: 10px; font-size: 13px;")
        layout.addWidget(self.lbl_resumen)
        
        # Tabla Unificada
        self.tabla = QTableWidget(0, 5)
        self.tabla.setHorizontalHeaderLabels(["Fila", "Código", "Descripción", "Estado", "Mensaje / Detalle"])
        self.tabla.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        self.tabla.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.tabla.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.tabla.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.tabla)
        
        # Boton Confirmar
        self.btn_confirmar = QPushButton("✅ Confirmar Importación Segura")
        self.btn_confirmar.setStyleSheet(f"background-color: {COLOR_SUCCESS}; color: white; padding: 12px; border-radius: 4px; font-weight: bold; font-size: 14px;")
        self.btn_confirmar.clicked.connect(self.ejecutar_importacion)
        self.btn_confirmar.setEnabled(False)
        layout.addWidget(self.btn_confirmar)

    def seleccionar_archivo(self):
        ruta, _ = QFileDialog.getOpenFileName(self, "Seleccionar Excel de Importación", "", "Excel Files (*.xlsx)")
        if not ruta:
            return
            
        self.lbl_archivo.setText(ruta)
        self.procesar_archivo(ruta)

    def auto_mapear_columnas(self, headers):
        mapa = {'codigo': -1, 'descripcion': -1, 'unidad': -1, 'precio_venta': -1, 'stock_inicial': -1, 'stock_minimo': -1, 'imagen': -1}
        aliases = {
            'codigo': ['código (*)', 'codigo', 'código', 'sku'],
            'descripcion': ['descripción (*)', 'descripción', 'descripcion', 'producto'],
            'unidad': ['unidad (*)', 'unidad', 'medida'],
            'precio_venta': ['precio venta (*)', 'precio venta', 'precio', 'precio_venta'],
            'stock_inicial': ['stock inicial', 'stock', 'existencia', 'cantidad'],
            'stock_minimo': ['stock mínimo', 'stock minimo', 'mínimo'],
            'imagen': ['imagen', 'img', 'imagen_path']
        }
        
        for idx, h in enumerate(headers):
            if not h: continue
            hl = str(h).strip().lower()
            for key, val_list in aliases.items():
                if hl in val_list and mapa[key] == -1:
                    mapa[key] = idx
                    break
        return mapa

    def procesar_archivo(self, ruta):
        self.filas_procesadas = []
        self.btn_confirmar.setEnabled(False)
        self.tabla.setRowCount(0)
        self.stats = {'total': 0, 'nuevos': 0, 'existentes': 0, 'errores': 0, 'ignoradas': 0}
        
        try:
            wb = openpyxl.load_workbook(ruta, data_only=True)
            ws = wb.active
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo leer el archivo Excel: {e}")
            return
            
        filas = list(ws.iter_rows(values_only=True))
        if not filas or len(filas) < 2:
            QMessageBox.warning(self, "Error", "El archivo está vacío o no tiene datos válidos.")
            return
            
        headers = filas[0]
        mapa = self.auto_mapear_columnas(headers)
        
        if mapa['codigo'] == -1 or mapa['descripcion'] == -1:
            QMessageBox.critical(self, "Formato No Reconocido", "No se encontraron las columnas 'Código' y 'Descripción'.")
            return
            
        from db.queries_stock import obtener_productos_activos_exportacion
        existentes = obtener_productos_activos_exportacion(self.conn)
        codigos_db = set(str(r[0]).strip() for r in existentes)
        
        codigos_excel = set()
        
        for i, fila in enumerate(filas[1:], start=2):
            if not any(fila):
                self.stats['ignoradas'] += 1
                continue
                
            self.stats['total'] += 1
            es_error = False
            estado = ""
            mensaje = ""
            
            try:
                cod = str(fila[mapa['codigo']]).strip() if mapa['codigo'] != -1 and fila[mapa['codigo']] is not None else ""
                desc = str(fila[mapa['descripcion']]).strip() if mapa['descripcion'] != -1 and fila[mapa['descripcion']] is not None else ""
                uni_raw = str(fila[mapa['unidad']]).strip() if mapa['unidad'] != -1 and fila[mapa['unidad']] is not None else ""
                
                precio = 0.0
                if mapa['precio_venta'] != -1 and fila[mapa['precio_venta']] is not None:
                    precio = float(fila[mapa['precio_venta']])
                    
                stk_ini = 0.0
                if mapa['stock_inicial'] != -1 and fila[mapa['stock_inicial']] is not None:
                    stk_ini = float(fila[mapa['stock_inicial']])
                    
                stk_min = 0.0
                if mapa['stock_minimo'] != -1 and fila[mapa['stock_minimo']] is not None:
                    stk_min = float(fila[mapa['stock_minimo']])
                    
                img = ""
                if mapa['imagen'] != -1 and fila[mapa['imagen']] is not None:
                    img = str(fila[mapa['imagen']]).strip()
                    
            except ValueError:
                self.agregar_fila_tabla(i, "N/A", "N/A", "✕ Error", "Valores numéricos inválidos", "#fef2f2")
                self.stats['errores'] += 1
                continue
                
            errs = []
            if not cod: errs.append("Código vacío")
            if not desc: errs.append("Descripción vacía")
            if precio < 0 or stk_ini < 0 or stk_min < 0: errs.append("Valores negativos")
            
            uni = normalizar_unidad(uni_raw)
            if not uni: errs.append(f"Unidad '{uni_raw}' inválida")
            
            if cod in codigos_excel: 
                errs.append("Código duplicado en Excel")
                
            if errs:
                mensaje = " | ".join(errs)
                self.agregar_fila_tabla(i, cod, desc, "✕ Error", mensaje, "#fef2f2")
                self.stats['errores'] += 1
                continue
                
            codigos_excel.add(cod)
            
            if cod in codigos_db:
                estado = "ACTUALIZAR"
                mensaje = "Se actualizarán datos maestros. Stock inicial ignorado."
                self.agregar_fila_tabla(i, cod, desc, "⚠ Advertencia", mensaje, "#fef08a")
                self.stats['existentes'] += 1
            else:
                estado = "NUEVO"
                mensaje = "Se creará producto. "
                if stk_ini > 0:
                    mensaje += f"Generará movimiento de entrada por {stk_ini}."
                self.agregar_fila_tabla(i, cod, desc, "✓ Válido", mensaje, "#f0fdf4")
                self.stats['nuevos'] += 1
                
            self.filas_procesadas.append((cod, desc, uni, precio, stk_ini, stk_min, img, estado))
            
        self.actualizar_resumen()

    def agregar_fila_tabla(self, fila_num, cod, desc, estado, msg, bg_color):
        idx = self.tabla.rowCount()
        self.tabla.insertRow(idx)
        
        items = [
            QTableWidgetItem(str(fila_num)),
            QTableWidgetItem(cod),
            QTableWidgetItem(desc),
            QTableWidgetItem(estado),
            QTableWidgetItem(msg)
        ]
        
        for it in items:
            it.setBackground(QColor(bg_color))
            
        self.tabla.setItem(idx, 0, items[0])
        self.tabla.setItem(idx, 1, items[1])
        self.tabla.setItem(idx, 2, items[2])
        self.tabla.setItem(idx, 3, items[3])
        self.tabla.setItem(idx, 4, items[4])

    def actualizar_resumen(self):
        st = self.stats
        resumen = (f"Total filas: {st['total']} | "
                   f"Nuevos: {st['nuevos']} | "
                   f"Existentes (Actualizar): {st['existentes']} | "
                   f"Errores: {st['errores']} | "
                   f"Ignoradas: {st['ignoradas']}")
                   
        if st['errores'] > 0:
            self.lbl_resumen.setText(f"❌ {resumen}. Hay errores que impedirán la importación.")
            self.lbl_resumen.setStyleSheet(f"color: {COLOR_DANGER}; font-weight: bold; margin-top: 10px; font-size: 13px;")
            self.btn_confirmar.setEnabled(False)
        else:
            self.lbl_resumen.setText(f"✅ {resumen}. Listo para importar.")
            self.lbl_resumen.setStyleSheet(f"color: {COLOR_SUCCESS}; font-weight: bold; margin-top: 10px; font-size: 13px;")
            if st['total'] > 0:
                self.btn_confirmar.setEnabled(True)

    def ejecutar_importacion(self):
        if not self.filas_procesadas or self.stats['errores'] > 0: return
        
        try:
            from db.queries_stock import ejecutar_importacion_excel_segura
            ejecutar_importacion_excel_segura(self.conn, self.filas_procesadas)
            QMessageBox.information(self, "Éxito", f"Se procesaron {len(self.filas_procesadas)} productos exitosamente.")
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Error Fatal", f"La importación falló y se realizó un ROLLBACK. Detalles:\n{e}")
